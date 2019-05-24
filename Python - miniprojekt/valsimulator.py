import openpyxl
import ui  # gränssnitt för terminalen
import random
import pandas

on = True  
vald_sida = "s"  
partier_info = []
klar = False
antal_röstande = 0
partier_invalda = []
antal_röster_invalda_partier = 0
procent_röster = 0
röstdeltagande = 0
partiledare = ["Isabella Löwin", "Jonas Sjöstedt", "Ulf Kristersson", "Stefan Löfven",
               "Jimmie Åkesson", "Annie Lööf","Jan Björklund", "Ebba Busch Thor",]
vidare = False
Block = ["Konservativa blocket", "Rödgröna"]
Inriktningar = ["Vänster", "Höger"]
Block_info = []
Inriktning_info = []
glada_partiledare = ""


def start(): #huvudet för gränssnittet i varje meny
    ui.line()
    ui.header("Val 2022")
    ui.line()
    ui.echo("En valsimulator inför valet 2022")
    ui.line(dots=True)


def slumpa(minimum, maximum):  # funktion för att slumpa mellan givna max och min värden
    slumptal = random.randint(minimum, maximum)
    return slumptal

df = pandas.read_csv('https://raw.githubusercontent.com/hjnilsson/SwedishPolls/master/Data/Polls.csv') #hämtar information med opinionsundersökningar från API och sparar i tabell

wb = openpyxl.load_workbook('valsim.xlsx') #öppnar excelfilen

sheet = wb.get_active_sheet() #använder den aktiva sidan då excelfilen bara innehåller en sida

# hämtar all info från exceldokumentet
partier_info = [{'Partinamn': sheet.cell(row=i, column=1).value,
                 'Inriktning': sheet.cell(row=i, column=2).value,
                 'Block': sheet.cell(row=i, column=3).value,
                 'Minimum_röst': sheet.cell(row=i, column=4).value,
                 'Maximum_röst': sheet.cell(row=i, column=5).value,
                 'Förkortning': sheet.cell(row=i, column=6).value,
                 }
                for i in range(2, sheet.max_row+1)]

for i in range(0, len(partier_info)):
    partier_info[i]['Partiledare'] = partiledare[i] #lägger till partiledare för varje parti

while on: #körs så länge on inte sätts till false
    if vald_sida == "x": #om användaren vill sluta programmmet rensas terminalen och whileloopen upphör
        ui.clear()
        on = False

    elif vald_sida == "k":
        ui.clear()
        start()
        ui.echo("Kommentarer:")

        for parti in partier_info:
            vidare = False #nollställer vidare för varje parti
            for parti_vidare in partier_invalda:
                if parti['Partinamn'] == parti_vidare['Partinamn']: #kollar ifall partiet matchar med det som gick vidare
                    vidare = True
            if vidare == False: # om ingen match gjordes är vidare fortfarande lika med false
                ui.echo(("Partiledaren " + parti['Partiledare'] + " är bedrövad över att " +
                         parti['Partinamn'] + " inte tog sig in i riksdagen.")) 
        
        partier_invalda_maxröst = partier_invalda #gör kopia av listan för enbart denna sidan så att den ordinarie listan inte påverkas
        partier_invalda_maxröst = sorted(
                partier_invalda_maxröst, key=lambda k: k['Andel_maxresultat'], reverse=True) #sorterar efter hur stor andel av det maximala resultatet som partierna fick
        
        ui.echo(glada_partiledare + " är glada över att vara en del av det största blocket.") #de partiledare som tillhör det vinnande blocket skrivs ut i meddelandet 

        ui.echo(partier_invalda_maxröst[0]['Partiledare'] + " är väldigt nöjd med det strålande valresultatet för " + partier_invalda_maxröst[0]['Partinamn'] + ".") 
        ui.echo(partier_invalda_maxröst[-1]['Partiledare'] + " är missnöjd med bottenresultatet för " + partier_invalda_maxröst[-1]['Partinamn'] + ".")
        ui.echo(" H - Hem")
        ui.echo(" X - Avsluta")
        vald_sida = ui.prompt(" Välj någon av ovan här:").lower() # tar input och ändrar vald sida 

    elif vald_sida == "o":
        ui.clear()
        start()
        ui.echo("Senaste opinionsundersökningen:")
        ui.echo("Institut: " + df['Company'][0]) # skriver ut första institutet i kolumnen Company, vilket är det senaste resultatet
        ui.echo("Undersökningen är gjord " + df['PublDate'][0] + "." )
        ui.echo("Moderaterna: " + str(df['M'][0]) + "%")
        ui.echo("Liberalerna: " + str(df['L'][0]) + "%")
        ui.echo("Centerpartiet: " + str(df['C'][0]) + "%")
        ui.echo("Kristdemokraterna: " + str(df['KD'][0]) + "%")
        ui.echo("Socialdemokraterna: " + str(df['S'][0]) + "%")
        ui.echo("Vänsterpartiet: " + str(df['V'][0]) + "%")
        ui.echo("Miljöpartiet: " + str(df['MP'][0]) + "%")
        ui.echo("Sverigedemokraterna: " + str(df['SD'][0]) + "%")
        ui.line()
        ui.echo(" H - Hem")
        ui.echo(" X - Avsluta")
        vald_sida = ui.prompt(" Välj någon av ovan här:").lower()

    elif vald_sida == "s" or vald_sida == "h":
        ui.clear()
        start()
        if vald_sida == "h":
            klar = True  # om användaren enbart vill tillbaka till hemsidan behöver ingen ny simulering ske
        else:
            klar = False
        while klar == False: # all kod som körs vid ny slumpnning. Även kod för de andra sidorna. Fördelen är att ingen kod behöver köras flera gånger i onödan.
            
            
            # nollställer all information från förra simuleringen
            antal_röstande = 0
            partier_invalda = []
            antal_röster_invalda_partier = 0
            procent_röster = 0
            röstdeltagande = 0
            klar = False
            Block_info = []
            Inriktning_info = []
            glada_partiledare = ""
            opinion = 0

            for parti in partier_info:
                # kör slumpfunktionen för varje parti
                parti['Röster'] = (slumpa(
                    parti['Minimum_röst'], parti['Maximum_röst']) + df[parti['Förkortning']][0])/2 #varje parti får antal röster i form av ett medelvärde mellan slumpningen och det senaste opinionsresultatet för partiet.
                
                antal_röstande += parti['Röster']

            # spärr för antal röster som krävs(4%)
            procentsspärr = antal_röstande * 0.04

            # skapar ny lista med alla partier som kom över procentspärren
            for parti in partier_info:
                if parti['Röster'] > procentsspärr:
                    partier_invalda.append(parti)

            # räknar röster för de invalda partierna
            for parti in partier_invalda:
                antal_röster_invalda_partier += parti['Röster']
            
            Block_info = [{'Block': block, #skapar lista med de block som angivits och tomma värdepar till senare kod
                'Storlek' : 0,
                'Partiledare' : [],
            }
            for block in Block]

            Inriktning_info = [{'Inriktning': inriktning,
                                'Storlek': 0,
                                'Partiledare' : [],
            }
            for inriktning in Inriktningar]

            
            for parti in partier_invalda:
                procent_röster = int(parti['Röster'] / antal_röster_invalda_partier * 100) # ger varje parti röster i procent som heltal
                parti['Procent_röster'] = procent_röster 

                parti['Andel_maxresultat'] = parti['Röster'] / parti['Maximum_röst'] #räknar ut andelen resultat jämfört med det bästa möjliga resultatet

                for block in Block_info:
                    if parti['Block'] == block['Block']:
                        block['Storlek'] += parti['Procent_röster'] # lägger till partiets storlek till rätt block
                        block['Partiledare'].append(parti['Partiledare']) 
                
                for inriktning in Inriktning_info:
                    if parti['Inriktning'] == inriktning['Inriktning']:
                        inriktning['Storlek'] += parti['Procent_röster']
                        inriktning['Partiledare'].append(parti['Partiledare'])

            
            partier_invalda = sorted(
                partier_invalda, key=lambda k: k['Procent_röster'], reverse=True)  # sorterar listan i storleksordning

            Block_info = sorted(
                Block_info, key=lambda k: k['Storlek'], reverse=True)

            Inriktning_info = sorted(
                Inriktning_info, key=lambda k: k['Storlek'], reverse=True)

            röstdeltagande = slumpa(50, 110) 

            if röstdeltagande <= 100:
                ui.echo(str(röstdeltagande) + " procent röstade i valet.")
                klar = True  # avslutar while-loopen då valresultatet är klart
            else:
                # kör while-loopen tills valdeltagandet är 100 eller mindre.
                ui.echo("Mer än 100 procent röstade. Ett omval hålls.")

            for i in range(0, len(Block_info[0]['Partiledare'])): #skapar del av meddelande med de partiledare som var nöjda med valresultatet
                if i == len(Block_info[0]['Partiledare']) - 2:
                    glada_partiledare += Block_info[0]['Partiledare'][i] + " och " # lägger till "och" efter den näst sista partiledaren
                elif i != len(Block_info[0]['Partiledare']) - 1:
                    glada_partiledare += Block_info[0]['Partiledare'][i] + ", " #lägger till "," om det inte är sista eller näst sista partiledaren
                else: 
                    glada_partiledare += Block_info[0]['Partiledare'][i] #lägger inte till något efter sista partiledaren

        ui.line()
        ui.echo("Valresultatet:")
        for parti in partier_invalda: #varje partis resultat
            ui.echo(parti['Partinamn'] + ": " +
                    str(parti['Procent_röster']) + "%")

        ui.echo(partier_invalda[0]['Partinamn'] + " är det störta partiet med " + str( #största partiet
            partier_invalda[0]['Procent_röster']) + " procent av rösterna.")

        ui.echo(Block_info[0]['Block'] + " är det största blocket med "+ str(Block_info[0]['Storlek']) + " procent av rösterna.")

        ui.echo("En majoritet " + Inriktning_info[0]['Inriktning'] + "orienterade blev invalda i riksdagen och fick totalt " + str(Inriktning_info[0]['Storlek']) + " av rösterna.")

        ui.line()
        ui.echo(" S - Simulera nytt valresultat")
        ui.echo(" K - Partiledarkommentarer")
        ui.echo(" O - Senaste opinionsundersökningen")
        ui.echo(" X - Avsluta")
        vald_sida = ui.prompt(" Välj någon av ovan här:").lower()
